from __future__ import annotations

from collections import Counter, defaultdict
from datetime import date, timedelta
from statistics import mean
import random
import string

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


TODAY = date(2026, 4, 18)
RANDOM_SEED = 2706
OUTPUT_FILE = "organization_structure_database.xlsx"


FIRST_NAMES = [
    "Aarav",
    "Olivia",
    "Liam",
    "Emma",
    "Noah",
    "Ava",
    "Sophia",
    "Isabella",
    "Mason",
    "Mia",
    "Lucas",
    "Amelia",
    "Ethan",
    "Harper",
    "James",
    "Evelyn",
    "Benjamin",
    "Abigail",
    "Elijah",
    "Emily",
    "Alexander",
    "Ella",
    "Michael",
    "Avery",
    "Daniel",
    "Sofia",
    "Henry",
    "Camila",
    "Sebastian",
    "Aria",
    "David",
    "Scarlett",
    "Joseph",
    "Victoria",
    "Samuel",
    "Madison",
    "Logan",
    "Luna",
    "Owen",
    "Grace",
    "Wyatt",
    "Chloe",
    "Luke",
    "Penelope",
    "Jack",
    "Layla",
    "Levi",
    "Riley",
    "Mateo",
    "Zoey",
    "Haruto",
    "Yuna",
    "Santiago",
    "Valentina",
    "Theo",
    "Nora",
    "Ananya",
    "Ishaan",
    "Fatima",
    "Amina",
    "Hassan",
    "Zara",
    "Aisha",
    "Rohan",
    "Anika",
    "Kabir",
    "Meera",
    "Nikhil",
    "Priya",
    "Aditya",
    "Kavya",
    "Siya",
    "Maya",
    "Yusuf",
    "Leila",
    "Chen",
    "Xinyi",
    "Jun",
    "Mei",
    "Hiro",
    "Kenji",
    "Sora",
    "Nina",
    "Arthur",
    "Matteo",
    "Chiara",
    "Luca",
    "Sven",
    "Greta",
    "Jonas",
    "Freya",
    "Ines",
    "Rafael",
    "Mariana",
    "Diego",
    "Helena",
    "Thiago",
    "Clara",
]


LAST_NAMES = [
    "Sharma",
    "Patel",
    "Singh",
    "Khan",
    "Smith",
    "Johnson",
    "Brown",
    "Taylor",
    "Anderson",
    "Thomas",
    "Lee",
    "Wilson",
    "Martin",
    "White",
    "Hall",
    "Adams",
    "Clark",
    "Lewis",
    "Walker",
    "Young",
    "Allen",
    "Wright",
    "King",
    "Scott",
    "Green",
    "Baker",
    "Nelson",
    "Carter",
    "Mitchell",
    "Perez",
    "Roberts",
    "Turner",
    "Phillips",
    "Campbell",
    "Parker",
    "Evans",
    "Edwards",
    "Collins",
    "Stewart",
    "Sanchez",
    "Morris",
    "Rogers",
    "Reed",
    "Cook",
    "Morgan",
    "Bell",
    "Murphy",
    "Bailey",
    "Rivera",
    "Cooper",
    "Richardson",
    "Cox",
    "Ward",
    "Torres",
    "Peterson",
    "Gray",
    "Ramirez",
    "James",
    "Watson",
    "Brooks",
    "Kelly",
    "Sanders",
    "Price",
    "Bennett",
    "Wood",
    "Barnes",
    "Ross",
    "Henderson",
    "Coleman",
    "Jenkins",
    "Perry",
    "Powell",
    "Long",
    "Flores",
    "Kim",
    "Tanaka",
    "Sato",
    "Watanabe",
    "Yamamoto",
    "Dubois",
    "Moreau",
    "Muller",
    "Schmidt",
    "Fischer",
    "Silva",
    "Santos",
    "Oliveira",
    "Costa",
    "Rossi",
    "Bianchi",
    "Romano",
    "Garcia",
    "Lopez",
    "Gonzalez",
    "Martinez",
    "Fernandez",
]


COUNTRY_META = {
    "USA": {
        "cities": [("New York", "New York"), ("Austin", "Texas"), ("Seattle", "Washington"), ("Chicago", "Illinois")],
        "region": "North America",
        "currency": "USD",
        "fx_to_usd": 1.00,
        "multiplier": 1.00,
        "phone_code": "+1",
        "legal_entity": "GlobalNova USA LLC",
    },
    "Canada": {
        "cities": [("Toronto", "Ontario"), ("Vancouver", "British Columbia"), ("Montreal", "Quebec")],
        "region": "North America",
        "currency": "CAD",
        "fx_to_usd": 1.35,
        "multiplier": 0.88,
        "phone_code": "+1",
        "legal_entity": "GlobalNova Canada Ltd",
    },
    "UK": {
        "cities": [("London", "England"), ("Manchester", "England"), ("Birmingham", "England")],
        "region": "EMEA",
        "currency": "GBP",
        "fx_to_usd": 0.79,
        "multiplier": 0.93,
        "phone_code": "+44",
        "legal_entity": "GlobalNova UK Ltd",
    },
    "Germany": {
        "cities": [("Berlin", "Berlin"), ("Munich", "Bavaria"), ("Frankfurt", "Hesse")],
        "region": "EMEA",
        "currency": "EUR",
        "fx_to_usd": 0.92,
        "multiplier": 0.95,
        "phone_code": "+49",
        "legal_entity": "GlobalNova Deutschland GmbH",
    },
    "India": {
        "cities": [("Bengaluru", "Karnataka"), ("Hyderabad", "Telangana"), ("Pune", "Maharashtra"), ("Gurgaon", "Haryana")],
        "region": "APAC",
        "currency": "INR",
        "fx_to_usd": 83.00,
        "multiplier": 0.42,
        "phone_code": "+91",
        "legal_entity": "GlobalNova India Pvt Ltd",
    },
    "Singapore": {
        "cities": [("Singapore", "Singapore")],
        "region": "APAC",
        "currency": "SGD",
        "fx_to_usd": 1.35,
        "multiplier": 0.90,
        "phone_code": "+65",
        "legal_entity": "GlobalNova Singapore Pte Ltd",
    },
    "Australia": {
        "cities": [("Sydney", "New South Wales"), ("Melbourne", "Victoria"), ("Brisbane", "Queensland")],
        "region": "APAC",
        "currency": "AUD",
        "fx_to_usd": 1.50,
        "multiplier": 0.94,
        "phone_code": "+61",
        "legal_entity": "GlobalNova Australia Pty Ltd",
    },
    "Brazil": {
        "cities": [("Sao Paulo", "Sao Paulo"), ("Rio de Janeiro", "Rio de Janeiro"), ("Campinas", "Sao Paulo")],
        "region": "LATAM",
        "currency": "BRL",
        "fx_to_usd": 5.10,
        "multiplier": 0.52,
        "phone_code": "+55",
        "legal_entity": "GlobalNova Brasil Ltda",
    },
    "UAE": {
        "cities": [("Dubai", "Dubai"), ("Abu Dhabi", "Abu Dhabi")],
        "region": "EMEA",
        "currency": "AED",
        "fx_to_usd": 3.67,
        "multiplier": 0.78,
        "phone_code": "+971",
        "legal_entity": "GlobalNova Middle East LLC",
    },
    "South Africa": {
        "cities": [("Johannesburg", "Gauteng"), ("Cape Town", "Western Cape"), ("Durban", "KwaZulu-Natal")],
        "region": "EMEA",
        "currency": "ZAR",
        "fx_to_usd": 18.20,
        "multiplier": 0.48,
        "phone_code": "+27",
        "legal_entity": "GlobalNova South Africa Pty Ltd",
    },
}


COUNTRY_QUOTAS = {
    "USA": 170,
    "India": 190,
    "UK": 95,
    "Germany": 90,
    "Canada": 90,
    "Singapore": 85,
    "Australia": 80,
    "Brazil": 80,
    "UAE": 65,
    "South Africa": 55,
}


FUNCTIONS = {
    "Engineering": {
        "code": "ENG",
        "business_unit": "Digital Products",
        "sub_functions": ["Platform Engineering", "Application Engineering", "Infrastructure Engineering"],
        "ic_titles": ["Software Engineer", "Senior Software Engineer", "Staff Software Engineer"],
        "manager_title": "Engineering Manager",
        "preferred_countries": ["India", "USA", "Germany", "Canada", "Singapore"],
        "ic_density": "high",
    },
    "Product Management": {
        "code": "PRD",
        "business_unit": "Digital Products",
        "sub_functions": ["Core Product", "Growth Product", "Product Operations"],
        "ic_titles": ["Product Analyst", "Product Manager", "Senior Product Manager"],
        "manager_title": "Product Manager, Team Lead",
        "preferred_countries": ["USA", "UK", "India", "Singapore", "Canada"],
        "ic_density": "medium",
    },
    "Sales": {
        "code": "SAL",
        "business_unit": "Enterprise Solutions",
        "sub_functions": ["Enterprise Sales", "Mid-Market Sales", "Inside Sales"],
        "ic_titles": ["Account Executive", "Senior Account Executive", "Sales Specialist"],
        "manager_title": "Sales Manager",
        "preferred_countries": ["USA", "UK", "Germany", "Brazil", "UAE", "Australia"],
        "ic_density": "high",
    },
    "Marketing": {
        "code": "MKT",
        "business_unit": "Enterprise Solutions",
        "sub_functions": ["Brand Marketing", "Performance Marketing", "Content Marketing"],
        "ic_titles": ["Marketing Specialist", "Senior Marketing Specialist", "Content Strategist"],
        "manager_title": "Marketing Manager",
        "preferred_countries": ["USA", "UK", "Canada", "Australia", "Brazil"],
        "ic_density": "medium",
    },
    "Finance": {
        "code": "FIN",
        "business_unit": "Corporate",
        "sub_functions": ["Accounting", "FP&A", "Treasury"],
        "ic_titles": ["Financial Analyst", "Senior Financial Analyst", "Accountant"],
        "manager_title": "Finance Manager",
        "preferred_countries": ["USA", "UK", "India", "Singapore", "Germany"],
        "ic_density": "medium",
    },
    "Human Resources": {
        "code": "HR",
        "business_unit": "Corporate",
        "sub_functions": ["Talent Acquisition", "People Operations", "Learning and Development"],
        "ic_titles": ["HR Generalist", "Talent Acquisition Specialist", "People Operations Specialist"],
        "manager_title": "HR Manager",
        "preferred_countries": ["USA", "India", "UK", "Canada", "South Africa"],
        "ic_density": "low",
    },
    "Operations": {
        "code": "OPS",
        "business_unit": "Global Services",
        "sub_functions": ["Business Operations", "Program Management", "Facilities"],
        "ic_titles": ["Operations Analyst", "Program Coordinator", "Business Operations Specialist"],
        "manager_title": "Operations Manager",
        "preferred_countries": ["India", "USA", "Singapore", "UAE", "Brazil"],
        "ic_density": "medium",
    },
    "Customer Support": {
        "code": "CS",
        "business_unit": "Enterprise Solutions",
        "sub_functions": ["Technical Support", "Customer Success", "Support Operations"],
        "ic_titles": ["Support Specialist", "Senior Support Specialist", "Customer Success Analyst"],
        "manager_title": "Customer Support Manager",
        "preferred_countries": ["India", "South Africa", "Brazil", "Canada", "Australia"],
        "ic_density": "high",
    },
    "Legal and Compliance": {
        "code": "LEG",
        "business_unit": "Corporate",
        "sub_functions": ["Legal Counsel", "Regulatory Compliance", "Contract Management"],
        "ic_titles": ["Compliance Analyst", "Legal Analyst", "Contract Specialist"],
        "manager_title": "Legal and Compliance Manager",
        "preferred_countries": ["USA", "UK", "Germany", "UAE"],
        "ic_density": "low",
    },
    "Information Technology": {
        "code": "IT",
        "business_unit": "Global Services",
        "sub_functions": ["IT Infrastructure", "Enterprise Applications", "Security Operations"],
        "ic_titles": ["Systems Administrator", "IT Support Engineer", "Security Analyst"],
        "manager_title": "IT Manager",
        "preferred_countries": ["India", "USA", "Singapore", "Germany", "Canada"],
        "ic_density": "medium",
    },
    "Data and Analytics": {
        "code": "DATA",
        "business_unit": "Digital Products",
        "sub_functions": ["Business Intelligence", "Data Engineering", "Data Science"],
        "ic_titles": ["Data Analyst", "BI Analyst", "Data Engineer"],
        "manager_title": "Data and Analytics Manager",
        "preferred_countries": ["India", "USA", "Canada", "UK", "Germany"],
        "ic_density": "high",
    },
    "Procurement": {
        "code": "PRC",
        "business_unit": "Global Services",
        "sub_functions": ["Strategic Sourcing", "Vendor Management", "Procurement Operations"],
        "ic_titles": ["Procurement Analyst", "Sourcing Specialist", "Vendor Operations Analyst"],
        "manager_title": "Procurement Manager",
        "preferred_countries": ["India", "UAE", "Germany", "Brazil", "Singapore"],
        "ic_density": "low",
    },
    "Supply Chain": {
        "code": "SCM",
        "business_unit": "Global Services",
        "sub_functions": ["Demand Planning", "Logistics", "Inventory Management"],
        "ic_titles": ["Supply Chain Analyst", "Logistics Coordinator", "Inventory Planner"],
        "manager_title": "Supply Chain Manager",
        "preferred_countries": ["Germany", "Brazil", "UAE", "India", "USA"],
        "ic_density": "medium",
    },
    "Research and Development": {
        "code": "RND",
        "business_unit": "Digital Products",
        "sub_functions": ["Applied Research", "Innovation Lab", "Prototype Engineering"],
        "ic_titles": ["Research Associate", "R&D Engineer", "Innovation Analyst"],
        "manager_title": "R&D Manager",
        "preferred_countries": ["USA", "Germany", "India", "UK", "Canada"],
        "ic_density": "high",
    },
    "Quality Assurance": {
        "code": "QA",
        "business_unit": "Digital Products",
        "sub_functions": ["Test Engineering", "Automation QA", "Release Quality"],
        "ic_titles": ["QA Analyst", "Test Engineer", "Automation QA Engineer"],
        "manager_title": "QA Manager",
        "preferred_countries": ["India", "Canada", "Germany", "South Africa", "UK"],
        "ic_density": "high",
    },
    "Corporate Strategy": {
        "code": "STR",
        "business_unit": "Corporate",
        "sub_functions": ["Strategic Planning", "M&A Analysis", "Program Management Office"],
        "ic_titles": ["Strategy Analyst", "Business Strategy Associate", "PMO Analyst"],
        "manager_title": "Strategy Manager",
        "preferred_countries": ["USA", "UK", "Singapore", "UAE", "Germany"],
        "ic_density": "low",
    },
}


SALARY_USD_BY_LEVEL = {
    "CEO": 420000,
    "Function Head": 260000,
    "Director": 180000,
    "Manager": 125000,
    "Individual Contributor": 82000,
}


BONUS_TARGET_BY_LEVEL = {
    "CEO": 40,
    "Function Head": 30,
    "Director": 20,
    "Manager": 12,
    "Individual Contributor": 8,
}


GRADE_BY_LEVEL = {
    "CEO": "G12",
    "Function Head": "G11",
    "Director": "G9",
    "Manager": "G7",
    "Individual Contributor": "G5",
}


EDUCATION_BY_LEVEL = {
    "CEO": ["MBA", "Masters", "PhD"],
    "Function Head": ["MBA", "Masters", "PhD"],
    "Director": ["MBA", "Masters", "Bachelors"],
    "Manager": ["Masters", "Bachelors"],
    "Individual Contributor": ["Bachelors", "Masters", "Associate Degree"],
}


class CountryAllocator:
    def __init__(self, quotas: dict[str, int]) -> None:
        self.remaining = dict(quotas)

    def assign(self, preferred: list[str] | None = None, forced: str | None = None) -> str:
        if forced is not None:
            if self.remaining.get(forced, 0) <= 0:
                raise ValueError(f"Country quota exhausted for forced assignment: {forced}")
            self.remaining[forced] -= 1
            return forced

        candidate = None
        candidate_quota = -1
        if preferred:
            for country in preferred:
                quota = self.remaining.get(country, 0)
                if quota > candidate_quota and quota > 0:
                    candidate = country
                    candidate_quota = quota

        if candidate is None:
            for country, quota in self.remaining.items():
                if quota > candidate_quota and quota > 0:
                    candidate = country
                    candidate_quota = quota

        if candidate is None:
            raise ValueError("No quota remaining for any country")

        self.remaining[candidate] -= 1
        return candidate


def rand_date(start: date, end: date) -> date:
    span_days = (end - start).days
    return start + timedelta(days=random.randint(0, span_days))


def pick_age_band(level: str) -> tuple[int, int]:
    if level == "CEO":
        return (46, 60)
    if level == "Function Head":
        return (40, 57)
    if level == "Director":
        return (34, 53)
    if level == "Manager":
        return (29, 48)
    return (22, 45)


def create_phone(country: str) -> str:
    meta = COUNTRY_META[country]
    code = meta["phone_code"]
    digits = "".join(random.choice(string.digits) for _ in range(9))
    return f"{code}-{digits[:3]}-{digits[3:6]}-{digits[6:]}"


def pick_gender() -> str:
    return random.choices(["Female", "Male", "Non-Binary"], weights=[47, 50, 3], k=1)[0]


def calc_tenure(hire_date: date) -> float:
    return round((TODAY - hire_date).days / 365.25, 1)


def salary_for(level: str, country: str) -> tuple[int, int, str]:
    base = SALARY_USD_BY_LEVEL[level]
    multiplier = COUNTRY_META[country]["multiplier"]
    noisy_usd = int(round(base * multiplier * random.uniform(0.88, 1.18), -2))
    currency = COUNTRY_META[country]["currency"]
    fx = COUNTRY_META[country]["fx_to_usd"]
    local_amount = int(round(noisy_usd * fx, -2))
    return noisy_usd, local_amount, currency


def hire_window(level: str) -> tuple[date, date]:
    if level == "CEO":
        return date(2010, 1, 1), date(2014, 12, 31)
    if level == "Function Head":
        return date(2011, 1, 1), date(2018, 12, 31)
    if level == "Director":
        return date(2014, 1, 1), date(2021, 12, 31)
    if level == "Manager":
        return date(2016, 1, 1), date(2024, 6, 30)
    return date(2018, 1, 1), date(2026, 3, 31)


def create_employee(
    employee_id: str,
    function_name: str,
    sub_function: str,
    designation: str,
    level: str,
    org_level: int,
    reporting_manager_id: str,
    function_head_id: str,
    allocator: CountryAllocator,
    preferred_countries: list[str],
    forced_country: str | None = None,
) -> dict[str, object]:
    first_name = random.choice(FIRST_NAMES)
    last_name = random.choice(LAST_NAMES)
    full_name = f"{first_name} {last_name}"
    country = allocator.assign(preferred=preferred_countries, forced=forced_country)
    city, state = random.choice(COUNTRY_META[country]["cities"])
    hire_start, hire_end = hire_window(level)
    hire_date = rand_date(hire_start, hire_end)
    age_start, age_end = pick_age_band(level)
    age = random.randint(age_start, age_end)
    dob = TODAY - timedelta(days=age * 365 + random.randint(0, 364))
    tenure = calc_tenure(hire_date)
    bonus = BONUS_TARGET_BY_LEVEL[level]
    salary_usd, salary_local, currency = salary_for(level, country)
    grade = GRADE_BY_LEVEL[level]
    education = random.choice(EDUCATION_BY_LEVEL[level])
    employment_type = "Contract" if level == "Individual Contributor" and random.random() < 0.08 else "Full-Time"
    fte = 0.8 if employment_type == "Contract" else 1.0

    if level in {"CEO", "Function Head", "Director"}:
        work_mode = random.choices(["Onsite", "Hybrid"], weights=[35, 65], k=1)[0]
    elif level == "Manager":
        work_mode = random.choices(["Onsite", "Hybrid", "Remote"], weights=[20, 65, 15], k=1)[0]
    else:
        work_mode = random.choices(["Onsite", "Hybrid", "Remote"], weights=[20, 55, 25], k=1)[0]

    if level == "CEO":
        office_type = "Global HQ"
    elif country in {"Singapore", "UK", "UAE", "USA"}:
        office_type = "Regional Hub"
    else:
        office_type = "Delivery Center"

    last_promotion = rand_date(hire_date, TODAY) if tenure >= 1 else hire_date
    performance = round(random.uniform(3.0, 4.9), 1)

    critical_role = "Yes" if level in {"CEO", "Function Head", "Director"} else "No"
    if level in {"Manager", "Individual Contributor"} and function_name in {"Engineering", "Data and Analytics", "Research and Development"}:
        if random.random() < 0.22:
            critical_role = "Yes"

    function_code = FUNCTIONS[function_name]["code"] if function_name in FUNCTIONS else "EXE"
    region = COUNTRY_META[country]["region"]
    region_code = "".join(ch for ch in region if ch.isupper())
    if not region_code:
        region_code = region[:3].upper()
    cost_center = f"{function_code}-{region_code}-{random.randint(100, 999)}"

    email = f"{first_name.lower()}.{last_name.lower()}.{employee_id[-4:]}@globalnova.com"

    return {
        "Employee_ID": employee_id,
        "First_Name": first_name,
        "Last_Name": last_name,
        "Full_Name": full_name,
        "Gender": pick_gender(),
        "Date_of_Birth": dob.isoformat(),
        "Hire_Date": hire_date.isoformat(),
        "Tenure_Years": tenure,
        "Employment_Status": "Active",
        "Employment_Type": employment_type,
        "FTE": fte,
        "Designation": designation,
        "Job_Level": level,
        "Org_Level": org_level,
        "Function": function_name,
        "Sub_Function": sub_function,
        "Business_Unit": FUNCTIONS[function_name]["business_unit"] if function_name in FUNCTIONS else "Corporate",
        "Reporting_Manager_ID": reporting_manager_id,
        "Function_Head_ID": function_head_id,
        "Is_People_Manager": "No",
        "Direct_Reports": 0,
        "Manager_Chain": "",
        "Work_Email": email,
        "Work_Phone": create_phone(country),
        "Location_City": city,
        "State_Province": state,
        "Country": country,
        "Region": region,
        "Work_Mode": work_mode,
        "Office_Type": office_type,
        "Legal_Entity": COUNTRY_META[country]["legal_entity"],
        "Cost_Center": cost_center,
        "Grade": grade,
        "Salary_Currency": currency,
        "Annual_Base_Salary": salary_local,
        "Annual_Base_Salary_USD": salary_usd,
        "Bonus_Target_Pct": bonus,
        "Last_Promotion_Date": last_promotion.isoformat(),
        "Performance_Rating_2025": performance,
        "Highest_Education": education,
        "Critical_Role": critical_role,
    }


def manager_team_size_band(function_name: str) -> int:
    density = FUNCTIONS[function_name]["ic_density"]
    if density == "high":
        return random.choice([3, 4, 5, 6])
    if density == "medium":
        return random.choice([3, 4, 5])
    return random.choice([2, 3, 4])


def build_dataset() -> list[dict[str, object]]:
    random.seed(RANDOM_SEED)
    allocator = CountryAllocator(COUNTRY_QUOTAS)
    employees: list[dict[str, object]] = []
    employee_counter = 1

    function_head_by_function: dict[str, str] = {}
    director_slots: list[tuple[str, str]] = []
    manager_slots: list[tuple[str, str]] = []

    def next_id() -> str:
        nonlocal employee_counter
        employee_id = f"EMP{employee_counter:04d}"
        employee_counter += 1
        return employee_id

    ceo_id = next_id()
    ceo = create_employee(
        employee_id=ceo_id,
        function_name="Corporate Strategy",
        sub_function="Executive Office",
        designation="Chief Executive Officer",
        level="CEO",
        org_level=1,
        reporting_manager_id="",
        function_head_id="",
        allocator=allocator,
        preferred_countries=["USA"],
        forced_country="USA",
    )
    ceo["Business_Unit"] = "Corporate"
    ceo["Function"] = "Executive Leadership"
    ceo["Sub_Function"] = "Executive Office"
    employees.append(ceo)

    # Level 2: Function heads
    for function_name, meta in FUNCTIONS.items():
        head_id = next_id()
        head = create_employee(
            employee_id=head_id,
            function_name=function_name,
            sub_function=random.choice(meta["sub_functions"]),
            designation=f"Vice President, {function_name}",
            level="Function Head",
            org_level=2,
            reporting_manager_id=ceo_id,
            function_head_id=head_id,
            allocator=allocator,
            preferred_countries=meta["preferred_countries"],
        )
        employees.append(head)
        function_head_by_function[function_name] = head_id

    # Level 3: Directors (3 per function)
    for function_name, meta in FUNCTIONS.items():
        for _ in range(3):
            director_id = next_id()
            director = create_employee(
                employee_id=director_id,
                function_name=function_name,
                sub_function=random.choice(meta["sub_functions"]),
                designation=f"Director, {function_name}",
                level="Director",
                org_level=3,
                reporting_manager_id=function_head_by_function[function_name],
                function_head_id=function_head_by_function[function_name],
                allocator=allocator,
                preferred_countries=meta["preferred_countries"],
            )
            employees.append(director)
            director_slots.append((director_id, function_name))

    # Level 4: Managers (4 per director)
    for director_id, function_name in director_slots:
        meta = FUNCTIONS[function_name]
        for _ in range(4):
            manager_id = next_id()
            manager = create_employee(
                employee_id=manager_id,
                function_name=function_name,
                sub_function=random.choice(meta["sub_functions"]),
                designation=meta["manager_title"],
                level="Manager",
                org_level=4,
                reporting_manager_id=director_id,
                function_head_id=function_head_by_function[function_name],
                allocator=allocator,
                preferred_countries=meta["preferred_countries"],
            )
            employees.append(manager)
            manager_slots.append((manager_id, function_name))

    # Level 5: Individual contributors to hit 1000 total employees.
    target_total = 1000
    current_total = len(employees)
    ic_target = target_total - current_total

    team_sizes = [manager_team_size_band(function_name) for _, function_name in manager_slots]
    while sum(team_sizes) < ic_target:
        idx = random.randrange(len(team_sizes))
        if team_sizes[idx] < 7:
            team_sizes[idx] += 1
    while sum(team_sizes) > ic_target:
        idx = random.randrange(len(team_sizes))
        if team_sizes[idx] > 2:
            team_sizes[idx] -= 1

    for (manager_id, function_name), team_size in zip(manager_slots, team_sizes):
        meta = FUNCTIONS[function_name]
        for _ in range(team_size):
            ic_id = next_id()
            designation = random.choice(meta["ic_titles"])
            individual = create_employee(
                employee_id=ic_id,
                function_name=function_name,
                sub_function=random.choice(meta["sub_functions"]),
                designation=designation,
                level="Individual Contributor",
                org_level=5,
                reporting_manager_id=manager_id,
                function_head_id=function_head_by_function[function_name],
                allocator=allocator,
                preferred_countries=meta["preferred_countries"],
            )
            employees.append(individual)

    if len(employees) != target_total:
        raise ValueError(f"Expected {target_total} employees, found {len(employees)}")

    if any(value != 0 for value in allocator.remaining.values()):
        raise ValueError(f"Country quotas were not fully allocated: {allocator.remaining}")

    by_id = {employee["Employee_ID"]: employee for employee in employees}
    direct_reports = defaultdict(int)
    for employee in employees:
        manager_id = employee["Reporting_Manager_ID"]
        if manager_id:
            direct_reports[manager_id] += 1

    manager_chain_cache: dict[str, list[str]] = {}

    def manager_chain(employee_id: str) -> list[str]:
        if employee_id in manager_chain_cache:
            return manager_chain_cache[employee_id]
        employee = by_id[employee_id]
        manager_id = employee["Reporting_Manager_ID"]
        if not manager_id:
            manager_chain_cache[employee_id] = []
            return []
        chain = manager_chain(manager_id) + [manager_id]
        manager_chain_cache[employee_id] = chain
        return chain

    for employee in employees:
        emp_id = employee["Employee_ID"]
        mgr_id = employee["Reporting_Manager_ID"]
        head_id = employee["Function_Head_ID"]
        employee["Direct_Reports"] = direct_reports[emp_id]
        employee["Is_People_Manager"] = "Yes" if direct_reports[emp_id] > 0 else "No"
        employee["Reporting_Manager_Name"] = by_id[mgr_id]["Full_Name"] if mgr_id else ""
        employee["Function_Head_Name"] = by_id[head_id]["Full_Name"] if head_id else ""
        chain_ids = manager_chain(emp_id)
        chain_names = [by_id[item]["Full_Name"] for item in chain_ids]
        employee["Manager_Chain"] = " > ".join(chain_names)

    return employees


def autosize_columns(worksheet) -> None:
    max_width = {}
    for row in worksheet.iter_rows(values_only=True):
        for idx, cell_value in enumerate(row, start=1):
            width = len(str(cell_value)) if cell_value is not None else 0
            max_width[idx] = min(max(max_width.get(idx, 0), width + 2), 40)
    for idx, width in max_width.items():
        worksheet.column_dimensions[get_column_letter(idx)].width = width


def write_excel(employees: list[dict[str, object]], file_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Employee_Master"

    columns = [
        "Employee_ID",
        "First_Name",
        "Last_Name",
        "Full_Name",
        "Gender",
        "Date_of_Birth",
        "Hire_Date",
        "Tenure_Years",
        "Employment_Status",
        "Employment_Type",
        "FTE",
        "Designation",
        "Job_Level",
        "Org_Level",
        "Function",
        "Sub_Function",
        "Business_Unit",
        "Reporting_Manager_ID",
        "Reporting_Manager_Name",
        "Function_Head_ID",
        "Function_Head_Name",
        "Is_People_Manager",
        "Direct_Reports",
        "Manager_Chain",
        "Work_Email",
        "Work_Phone",
        "Location_City",
        "State_Province",
        "Country",
        "Region",
        "Work_Mode",
        "Office_Type",
        "Legal_Entity",
        "Cost_Center",
        "Grade",
        "Salary_Currency",
        "Annual_Base_Salary",
        "Annual_Base_Salary_USD",
        "Bonus_Target_Pct",
        "Last_Promotion_Date",
        "Performance_Rating_2025",
        "Highest_Education",
        "Critical_Role",
    ]

    ws.append(columns)
    for employee in employees:
        ws.append([employee[column] for column in columns])

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    autosize_columns(ws)

    summary = wb.create_sheet("Org_Summary")
    summary["A1"] = "Metric"
    summary["B1"] = "Value"
    summary["A1"].font = Font(bold=True)
    summary["B1"].font = Font(bold=True)

    people_managers = sum(1 for employee in employees if employee["Is_People_Manager"] == "Yes")
    avg_tenure = round(mean(float(employee["Tenure_Years"]) for employee in employees), 2)
    summary_rows = [
        ("Generated_On", TODAY.isoformat()),
        ("Total_Employees", len(employees)),
        ("Total_Functions", len(FUNCTIONS)),
        ("Total_Countries", len(COUNTRY_QUOTAS)),
        ("People_Managers", people_managers),
        ("Average_Tenure_Years", avg_tenure),
    ]
    for row in summary_rows:
        summary.append(row)

    summary.append(("", ""))
    summary.append(("Headcount by Function", ""))
    summary["A9"].font = Font(bold=True)
    function_counts = Counter(employee["Function"] for employee in employees)
    for function_name, count in sorted(function_counts.items(), key=lambda item: item[0]):
        summary.append((function_name, count))

    summary.append(("", ""))
    summary.append(("Headcount by Country", ""))
    summary[f"A{summary.max_row}"].font = Font(bold=True)
    country_counts = Counter(employee["Country"] for employee in employees)
    for country_name, count in sorted(country_counts.items(), key=lambda item: item[0]):
        summary.append((country_name, count))

    summary.freeze_panes = "A2"
    autosize_columns(summary)
    wb.save(file_path)


def main() -> None:
    employees = build_dataset()
    write_excel(employees, OUTPUT_FILE)
    print(f"Created {OUTPUT_FILE} with {len(employees)} employees.")


if __name__ == "__main__":
    main()
